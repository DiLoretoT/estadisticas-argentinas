"""
Fetcher de deuda pública argentina.

Fuente: MECON Secretaría de Finanzas — Excel trimestral.
URL pattern: https://www.argentina.gob.ar/sites/default/files/deuda_publica_DD-MM-YYYY.xlsx

Auto-detecta el último archivo disponible probando trimestres en orden
descendente. El parser lee la hoja A.2.5 que tiene la serie histórica
anual del Sector Público Nacional desde 1992.

Series ingestadas (anuales, USD millones):
  - deuda_total: TOTAL DEUDA PÚBLICA BRUTA
  - deuda_titulos: TÍTULOS PÚBLICOS Y LETRAS DEL TESORO
  - deuda_prestamos: PRÉSTAMOS (todos los acreedores)
  - deuda_organismos_intl: ORGANISMOS INTERNACIONALES (FMI, BIRF, BID, etc)
  - deuda_organismos_oficiales: ORGANISMOS OFICIALES bilaterales
  - deuda_adelantos_bcra: ADELANTOS TRANSITORIOS BCRA
  - deuda_pago_diferido: DEUDA EN SITUACIÓN DE PAGO DIFERIDO

Frecuencia: anual (cierre 31-dic de cada año).
"""

from __future__ import annotations

import io
import logging
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

import certifi
import openpyxl
import requests

from common import now_iso, today_iso, write_json
from db import (
    finish_refresh_run,
    init_db,
    start_refresh_run,
    update_series_refresh_status,
    upsert_observations,
    upsert_series,
)
from series_store import write_series

logger = logging.getLogger(__name__)

_BASE_URL = "https://www.argentina.gob.ar/sites/default/files"
_FILENAME_PATTERN = "deuda_publica_{day:02d}-{month:02d}-{year}.xlsx"

# Hoja con la serie histórica anual 1992-actual
_SHEET = "A.2.5"

# Mapeo de filas de A.2.5 a series del ETL.
# (row_index, series_key, display_name)
_ROWS: list[tuple[int, str, str]] = [
    (17, "total", "Deuda Pública Bruta — Total"),
    (22, "titulos", "Deuda — Títulos públicos y letras del tesoro"),
    (24, "prestamos", "Deuda — Préstamos (todos los acreedores)"),
    (28, "organismos_intl", "Deuda — Organismos internacionales (FMI, BIRF, BID)"),
    (30, "organismos_oficiales", "Deuda — Organismos oficiales bilaterales"),
    (36, "adelantos_bcra", "Deuda — Adelantos transitorios BCRA"),
    (38, "pago_diferido", "Deuda — En situación de pago diferido"),
]

# Fila 12 tiene los headers de fecha (31/12/1992, 1993-12-31, ..., 2025-12-31)
_DATES_ROW = 12

# Primera columna con valores numéricos (1992)
_FIRST_DATA_COL = 2

# Trimestres en orden de preferencia (más reciente primero)
_QUARTERS = [(12, 31), (9, 30), (6, 30), (3, 31)]


def _http_get(url: str) -> bytes:
    allow_insecure = os.getenv("ALLOW_INSECURE_SSL") == "1"
    verify = False if allow_insecure else certifi.where()
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (compatible; estadisticas-argentinas-bot/1.0)"
        )
    }
    resp = requests.get(url, headers=headers, timeout=60, verify=verify)
    resp.raise_for_status()
    return resp.content


def find_latest_xlsx_url() -> str | None:
    """Prueba año actual y anterior con los 4 trimestres hasta encontrar 200."""
    today = date.today()
    for year in (today.year, today.year - 1):
        for month, day in _QUARTERS:
            fname = _FILENAME_PATTERN.format(day=day, month=month, year=year)
            url = f"{_BASE_URL}/{fname}"
            try:
                allow_insecure = os.getenv("ALLOW_INSECURE_SSL") == "1"
                verify = False if allow_insecure else certifi.where()
                resp = requests.head(url, timeout=15, verify=verify, allow_redirects=True)
                if resp.status_code == 200:
                    logger.info("Excel MECON encontrado: %s", url)
                    return url
            except requests.RequestException as exc:
                logger.debug("Probe %s falló: %s", url, exc)
                continue
    return None


def _parse_sheet(content: bytes) -> dict[str, list[tuple[date, float]]]:
    """Devuelve {series_key: [(date, value), ...]} a partir del Excel."""
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(content), data_only=True, read_only=True
    )
    if _SHEET not in wb.sheetnames:
        raise ValueError(f"Hoja '{_SHEET}' no encontrada en el Excel")

    ws = wb[_SHEET]
    rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))

    # Header de fechas (fila 12)
    dates_row = rows[_DATES_ROW - 1]
    dates: list[date | None] = []
    for cell in dates_row:
        if cell is None:
            dates.append(None)
            continue
        if isinstance(cell, datetime):
            dates.append(cell.date())
        elif isinstance(cell, str) and cell.strip():
            # Probar parsear "31/12/92 (*)" como diciembre 1992
            cleaned = cell.split(" ")[0].replace("(*)", "").strip()
            try:
                if "/" in cleaned:
                    parts = cleaned.split("/")
                    day, month, year = (
                        int(parts[0]),
                        int(parts[1]),
                        int(parts[2]),
                    )
                    if year < 100:
                        year += 1900 if year >= 50 else 2000
                    dates.append(date(year, month, day))
                else:
                    dates.append(None)
            except (ValueError, IndexError):
                dates.append(None)
        else:
            dates.append(None)

    out: dict[str, list[tuple[date, float]]] = {}
    for row_idx, series_key, _display in _ROWS:
        data_row = rows[row_idx - 1]
        points: list[tuple[date, float]] = []
        for col_idx, cell in enumerate(data_row):
            if col_idx >= len(dates):
                continue
            d = dates[col_idx]
            if d is None or cell is None:
                continue
            try:
                v = float(cell)
            except (TypeError, ValueError):
                continue
            points.append((d, v))
        # Ordenar por fecha
        points.sort(key=lambda p: p[0])
        out[series_key] = points
    return out


def _persist(
    series_key: str,
    display_name: str,
    points: list[tuple[date, float]],
    excel_url: str,
) -> dict[str, Any]:
    full_key = f"deuda_{series_key}"
    diario_file = f"deuda_{series_key}.json"
    summary_file = f"deuda_{series_key}_summary.json"

    upsert_series(
        series_id=full_key,
        display_name=display_name,
        source_name="MECON Sec. Finanzas",
        dataset="deuda_publica.xlsx hoja A.2.5",
        official=True,
        frequency="annual",
        unit="millones_USD",
        provider_series_id=f"A.2.5_R{series_key}",
    )
    run_id = start_refresh_run(full_key)

    try:
        write_series(diario_file, points)
        rows_upserted = upsert_observations(full_key, points)

        if not points:
            payload = {
                "updated_at": None,
                "period": None,
                "value": None,
                "yoy_change": None,
                "unit": "millones_USD",
                "source": {
                    "name": "MECON",
                    "dataset": excel_url,
                    "official": True,
                },
            }
            finish_refresh_run(run_id, "success", 0, None)
            update_series_refresh_status(full_key, "success", None, 0, None)
            if os.getenv("ETL_EXPORT_JSON", "1") == "1":
                write_json(summary_file, payload)
            return payload

        latest_date, latest_value = points[-1]
        yoy_change = None
        if len(points) >= 2:
            prev = points[-2][1]
            if prev:
                yoy_change = round(((latest_value - prev) / prev) * 100, 2)

        payload = {
            "updated_at": today_iso(),
            "updated_at_time": now_iso(),
            "period": latest_date.strftime("%Y-%m-%d"),
            "value": round(latest_value, 2),
            "yoy_change": yoy_change,
            "unit": "millones_USD",
            "source": {"name": "MECON", "dataset": excel_url, "official": True},
        }
        finish_refresh_run(run_id, "success", rows_upserted, None)
        update_series_refresh_status(
            full_key, "success", latest_date, len(points), None
        )
        if os.getenv("ETL_EXPORT_JSON", "1") == "1":
            write_json(summary_file, payload)
        logger.info("%s: %d puntos, último %s = USD %s M",
                    full_key, len(points), latest_date, latest_value)
        return payload

    except Exception as exc:
        logger.exception("Error persistiendo %s", series_key)
        finish_refresh_run(run_id, "error", 0, str(exc))
        update_series_refresh_status(full_key, "error", None, 0, str(exc))
        raise


def fetch_deuda() -> dict[str, Any]:
    init_db()
    url = find_latest_xlsx_url()
    if not url:
        logger.error("No se encontró ningún Excel de deuda en MECON.")
        return {}

    logger.info("Descargando %s", url)
    content = _http_get(url)
    parsed = _parse_sheet(content)

    results: dict[str, Any] = {}
    failed: list[str] = []
    for row_idx, series_key, display_name in _ROWS:
        try:
            points = parsed.get(series_key, [])
            results[series_key] = _persist(series_key, display_name, points, url)
        except Exception as exc:
            logger.error("Falló serie deuda %s: %s", series_key, exc)
            failed.append(series_key)

    if failed:
        logger.warning("Series con error: %s", ", ".join(failed))
    return results


def main() -> None:
    fetch_deuda()


if __name__ == "__main__":
    main()
